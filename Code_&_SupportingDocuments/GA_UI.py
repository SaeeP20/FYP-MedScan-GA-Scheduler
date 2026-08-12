# UI for the genetic algorithm scheduling system

import openpyxl as pxl
from openpyxl.worksheet.worksheet import Worksheet
import pandas as pd

def load_schedule(filename):
    """
    Loads the schedule Excel workbook and returns the active sheet.
    
    Raises FileNotFoundError if schedule.xlsx is not found.
    Raises ValueError if no active sheet is available.
    
    Args:
        filename (str): The name of the Excel file to load.

    Returns:
        tuple: (workbook, active_sheet)
    """
    try:
        wb = pxl.load_workbook(filename) # O(1) lookup time
    except FileNotFoundError:
        raise FileNotFoundError("File not found. Please ensure the file is in the correct location.")
    sheet = wb.active
    if sheet is None:
        raise ValueError("\nNo active sheet found in schedule.xlsx.")
    return wb, sheet

def save_schedule(schedule, filename):
    """
    Saves the current schedule to an Excel file.
    
    Args:
        schedule: The schedule tuple (workbook, sheet).
        filename (str): The name of the file to save to.
    """
    wb, sheet = schedule
    if wb is not None:
        wb.save(filename)
        print("\nSchedule saved successfully.")
    else:
        raise ValueError("No workbook to save.")
    
def load_patients(filename):
    """
    Loads patient data from patients.csv into a pandas DataFrame.
    
    Expects CSV with columns: patient_id, name, preferred_time, preferred_day, preferred_slot.
    
    Args:
        filename (str): The name of the CSV file to load.

    Returns:
        pd.DataFrame: Patient data.
    """
    try:
        return pd.read_csv(filename)
    except pd.errors.EmptyDataError:
        print("Warning: patient file is empty. Returning empty DataFrame.")
        return pd.DataFrame(columns=["patient_id", "name", "preferred_time", "preferred_day", "preferred_slot"])

def save_patients(df, filename):
    """
    Saves the patient DataFrame to a CSV file.
    
    Args:
        df (pd.DataFrame): Patient records to save.
        filename (str): The name of the file to save to.
    """
    df.to_csv(filename, index=False)
    print("Patient data saved successfully.")

def update_patients(schedule, patients_df, name_normalised) -> pd.DataFrame:
    """
    Handles patient input for preferred day and time, validates against schedule, and updates patients.csv.
    
    Args:
        schedule: Excel sheet object for validation.
        patients_df (pd.DataFrame): Current patient data.
        name_normalised (str): Normalized patient name.
    
    Returns:
        pd.DataFrame: Updated patient data.
    """
    if schedule is None:
        raise ValueError("Schedule is not loaded. Cannot validate preferences.")
    else:
        wb, sheet = schedule

    # ======== preferred day =========
    pref_day = input("Enter the first 3 letters of your preferred day (e.g. Mon, Tue, or Any): ").capitalize()

    valid_days = ["Mon", "Tue", "Wed", "Thu", "Fri"]

    if pref_day == "" or pref_day not in valid_days:
        print("Invalid day. No changes made.")
        return patients_df

    if pref_day == "Any" or pref_day == "":
        pref_day = None

    # ======== preferred time ========
    raw = input("Enter your preferred time: ").strip().lower().replace(" ", "")

    # Load valid times from Excel
    valid_times = []
    for r in range(2, sheet.max_row + 1):
        val = sheet.cell(row=r, column=1).value

        # Convert time to 12-hour format if it's a datetime object
        if hasattr(val, "strftime"):
            val = val.strftime("%I:%M")

        # For consistent formatting
        if isinstance(val, str) and len(val) == 4:  # e.g. "4:00"
            val = "0" + val

        valid_times.append(val)

    # Handle "any"
    if raw == "any" or raw == "":
        pref_time = None
    else:
        # Extract hour
        if "am" in raw or "pm" in raw:
            hour = int(raw.split(":")[0].replace("am", "").replace("pm", ""))
        elif ":" in raw:
            hour = int(raw.split(":")[0])
        else:
            hour = int(raw)

        # Convert to Excel's 12-hour format with leading zero
        pref_time = f"{hour:02d}:00"

        # Validate against schedule
        if pref_time not in valid_times:
            print("Invalid time. No changes made.")
            return patients_df

    # ======= calculate preferred_time index as per excel ========
    day_index = valid_days.index(pref_day) if pref_day is not None else None

    times = valid_times
    time_index = times.index(pref_time) if pref_time is not None else None

    num_times = len(times)

    if day_index is None or time_index is None:
        preferred_time_index = None
    else:
        preferred_time_index = day_index * num_times + time_index

    # ======= update or add patient =========
    if name_normalised in patients_df['name'].values:
        patients_df.loc[patients_df['name'] == name_normalised, 'preferred_time'] = pref_time
        patients_df.loc[patients_df['name'] == name_normalised, 'preferred_day'] = pref_day
        patients_df.loc[patients_df['name'] == name_normalised, 'preferred_slot'] = preferred_time_index
        print("Preference updated.")
    else:
        new_patient = pd.DataFrame({
            'patient_id': [len(patients_df)+1],
            'name': [name_normalised],
            'preferred_time': [pref_time],
            'preferred_day': [pref_day],
            'preferred_slot': [preferred_time_index]
        })
        patients_df = pd.concat([patients_df, new_patient], ignore_index=True)
        print(f"New patient {name_normalised} added.")
        
    return patients_df  

def write_schedule_to_excel(schedule, filename, genotype, patients_df):
    """
    Writes the genotype schedule to the Excel file.
    
    Args:
        schedule: Excel sheet object for validation.
        filename (str): The name of the Excel file to save.
        genotype (list): List of slot assignments.
        patients_df (pd.DataFrame): Patient data for names.
    """
    wb, sheet = schedule
    NUM_TIMES = sheet.max_row - 1

    for patient_idx, slot in enumerate(genotype):
        day_index = slot // NUM_TIMES
        time_index = slot % NUM_TIMES

        row = time_index + 2
        col = day_index + 2

        if col < 2:
            raise ValueError(f"Invalid day_index={day_index}. Attempted to write into time column.")

        name = patients_df.iloc[patient_idx]["name"]
        sheet.cell(row=row, column=col).value = name

    save_schedule(schedule, filename)

def main():
    """
    Main UI function for patient scheduling system.
    
    Handles user input for patient or employee roles.
    """
    schedule = load_schedule("schedule.xlsx")
    patients = load_patients("patients.csv")
    user_type = input("Hello, are you a patient or employee? \n").lower()

    if user_type not in ("patient", "employee"):
        print("Invalid user type. Please enter 'patient' or 'employee'. Exiting.")
        return

    # ======== employee flow ========
    elif user_type == "employee":
        print("Welcome, employee! Would you like to create a schedule? (yes/no)")
        create_schedule = input().lower()
        if create_schedule == "yes":
            import GA_firstPT as ga
            print("Creating schedule using genetic algorithm...")
            # Call the function to create a schedule
            new_schedule = ga.genetic_algo(t_size=2, cross_prob=0.5, mut_prob=0.5)
            write_schedule_to_excel(schedule, "schedule.xlsx", new_schedule["best_geno"], patients)
            print("Schedule created and saved to schedule.xlsx")
        else:
            print("No schedule created. Exiting.")
        return

    # ======== patient flow =======
    name = input("Welcome, patient! Please enter your full name: \n").strip()
    name_normalized = name.title()
    print("Loading your preferences...")
    if name_normalized in patients['name'].values:
        pref_time = patients.loc[patients['name'] == name_normalized, 'preferred_time'].values[0]
        pref_day = patients.loc[patients['name'] == name_normalized, 'preferred_day'].values[0]
        print(f"Your current preferred time is: {pref_day} at {pref_time}")
        if pd.isna(pref_time):
            print("You currently have no preferred time set.")    
        choice = input("Would you like to update it? (yes/no): ").lower()
        if choice != "yes":
            print("No changes made.")
            return
    else:
        print("No existing preferences found. Let's set them up!")

    updated_patients = update_patients(schedule, patients, name_normalized)
    save_patients(updated_patients, "patients.csv")

if __name__ == "__main__":
    main()
    